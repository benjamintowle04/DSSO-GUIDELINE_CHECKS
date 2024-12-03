import os
import tkinter as tk
from tkinter import ttk, messagebox, Toplevel
from client.schedulesourcecalls import getLocations, getScheduleNames
from utils.helperFunctions import getLocationNames, convert_to_12hr_format
from controllers.checkGuidelines import runGuidelineCheck
from openpyxl import load_workbook
from tkcalendar import DateEntry
from datetime import date, datetime
import sys


if getattr(sys, 'frozen', False):
    # Running as compiled executable
    base_path = sys._MEIPASS
else:
    # Running in a normal Python environment
    base_path = os.path.dirname(os.path.abspath(__file__))

callsheet_path = os.path.join(base_path, 'Callsheet.xlsx')
wb = load_workbook(callsheet_path)


#Populates call sheet with the employees that are missing the requirements

def generate_call_sheet(location, scheduleId):
    start_date = start_input.get_date()
    end_date = end_input.get_date()

    #All computations for missing guidelines run through this function
    data = runGuidelineCheck(location, scheduleId, start_date, end_date)

    #Only 1 worksheet in Callsheet.xlsx
    ws = wb.active

    #Refresh the page so that no old data is still in there
    clear_all_entries()

    #So the user knows that the program actually ran, the date and time display on the callsheet
    todays_date = date.today()
    current_time = datetime.now().strftime("%H:%M:%S")

    ws["A3"] = "Facility: " + location
    ws["A4"] = "Date: " + todays_date.strftime("%m/%d/%Y")
    ws["B4"] = "Time: " + convert_to_12hr_format(current_time)

    #Init headers to columns
    ws["A6"] = "Last"
    ws["B6"] = "First"
    ws["C6"] = "# Shifts"
    ws["D6"] = "Hours"
    ws["E6"] = "SUP?"
    ws["F6"] = "Missing Requirement(s)"
    ws["G6"] = "Notes"

    k = 7

    # Add Rows to new worksheet starting from row 7
    for item in data:
        ws[f"A{k}"] = item.last_name
        ws[f"B{k}"] = item.first_name
        ws[f"C{k}"] = item.shift_count
        ws[f"D{k}"] = item.hours
        ws[f"E{k}"] = item.supervisor
        ws[f"F{k}"] = item.missing_reqs
        ws[f"G{k}"] = item.notes
        k += 1

    wb.save("Callsheet.xlsx")
    print("FINISHED")


#Refresh all data in the callsheet
#Clear every cell in the worksheet
def clear_all_entries():
    ws = wb.active
    for row in ws.iter_rows(min_row=7):
        for cell in row:
            cell.value = None




# Set the dropdown menu input fields and their respective values
# Called each time the location is changed in the first dropdown to re-update the schedules
def initDropdowns():
    # Label and Entry widget for Facility Name
    ttk.Label(root, text="Facility Name:").grid(column=0, row=0, padx=10, pady=5, sticky=tk.W)
    facility_name = ttk.Combobox(root, textvariable=selected_facility_var, values=sorted(location_name_list), width=30)
    facility_name.bind('<<ComboboxSelected>>', select_location)
    facility_name.grid(column=1, row=0, padx=10, pady=5)

    # Label and Entry widget for Schedule Name
    ttk.Label(root, text="Schedule Name:").grid(column=0, row=1, padx=10, pady=5, sticky=tk.W)
    schedule_name = ttk.Combobox(root, textvariable=selected_schedule_var, values=schedule_name_list, width=30)
    schedule_name.grid(column=1, row=1, padx=10, pady=5)


#Triggers a function that retrieves the list of schedules when an option is changed
#Used to allow us to pass in the location's name as a parameter to the actual function
def select_location(event):
    selected_facility = selected_facility_var.get()
    on_location_change(selected_facility)


#Search function that will display the list of schedules for the location selected in the first input
#Re-initializes dropdown menus to display schedules specific to the selected location
def on_location_change(newLocation):
    global schedule_name_list  # Declare that we are using the global variable
    global schedule_id_list
    schedule_name_id_list = getScheduleNames(newLocation)
    schedule_name_list = schedule_name_id_list.get("Names")
    schedule_id_list = schedule_name_id_list.get("ScheduleIds")
    initDropdowns()
    print("Location Changed to " + newLocation)


# Function to be called when the OK button is pressed
# Triggers the subprocess that runs gridGenerator.py
# When clicked, program will generate a window displaying where the new grid is
# Also displays a window to show available empty shifts given their availability
def on_ok():
    # Print the values entered in the entry fields
    print(f"Facility Name: {selected_facility_var.get()}")
    print(f"Schedule Name: {selected_schedule_var.get()}")

    try:
        id_index = get_schedule_name_index(selected_schedule_var.get())
        scheduleId = schedule_id_list[id_index]

        # Retrieve the schedule ID number based on the user input
        generate_call_sheet(selected_facility_var.get(), scheduleId)

        show_file_path(os.path.join(os.getcwd(), "Callsheet.xlsx"))

    except Exception as e:
        message = (f"Failed to Generate Call Sheet: Please try running the program again. Double check that dates, "
                   f"schedules and facilities are correct")
        messagebox.showerror("Error", message)




# Function to create a new window displaying the file path
#Displays where the callsheet is to the end user on the gui
def show_file_path(file_path):
    # Create a new top-level window
    new_window = tk.Toplevel(root)
    new_window.title("Generated Callsheet")
    new_window.geometry("600x200")

    # Display the file path
    ttk.Label(new_window, text="Call sheet generated at:").pack(pady=10)

    # Display the file path in a larger white font
    file_path_label = tk.Label(new_window, text=file_path, fg="black", font=("Helvetica", 15, "bold"))
    file_path_label.pack(pady=10)

    # Button to close the new window and terminate the program
    ttk.Button(new_window, text="OK", command=root.quit).pack(pady=10)


#Initializes the visual date input for the end user
#Allows the end user to input their start and end dates for the schedule
def init_date_entry(root, row, label_text):
    label = tk.Label(root, text=label_text)
    label.grid(row=row, column=0, padx=10, pady=5)

    date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2)
    date_entry.grid(row=row, column=1, padx=5, pady=5)

    return date_entry

#Returns the index for the requested schedule in the list of active ones
#Used to retrieve the schedule ID that is associated with the schedule's name
def get_schedule_name_index(name):
    for index, item in enumerate(schedule_name_list):
        if item == name:
            return index
    return -1






# Create the main window
root = tk.Tk()
root.title("DSSO Guideline Check")
root.geometry("440x160")

#Create lists of locations for the first dropdown menu
location_list = getLocations()
location_name_list = getLocationNames(location_list)

#Keeps track of the location and schedule selected from the dropdown menu
selected_facility_var = tk.StringVar(root)
selected_schedule_var = tk.StringVar(root)

#Set the initital dropdown value
selected_facility_var.set("Select a location")


#Declare variables that will get initialized in initDropdowns()
schedule_name_list = []
schedule_id_list = []

initDropdowns()

#Initialize Date Range Entries
start_input = init_date_entry(root, 3, "Start Date")
end_input = init_date_entry(root, 4, "End Date")


#Initialize Values

# Button to generate the schedule
generate_button = ttk.Button(root, text="Generate Call Sheet", command=on_ok)
generate_button.grid(column=0, row=5, columnspan=2, pady=10)

# Add padding around the entire grid
for child in root.winfo_children():
    child.grid_configure(padx=10, pady=5)


# Run the application
root.mainloop()
